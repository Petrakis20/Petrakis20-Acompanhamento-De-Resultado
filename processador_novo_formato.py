import argparse
import io
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl
import xlrd


MESES_ALIASES = {
    "janeiro": "janeiro",
    "jan": "janeiro",
    "01": "janeiro",
    "fevereiro": "fevereiro",
    "fev": "fevereiro",
    "02": "fevereiro",
    "marco": "marco",
    "mar": "marco",
    "03": "marco",
    "abril": "abril",
    "abr": "abril",
    "04": "abril",
    "maio": "maio",
    "mai": "maio",
    "05": "maio",
    "junho": "junho",
    "jun": "junho",
    "06": "junho",
    "julho": "julho",
    "jul": "julho",
    "07": "julho",
    "agosto": "agosto",
    "ago": "agosto",
    "08": "agosto",
    "setembro": "setembro",
    "set": "setembro",
    "09": "setembro",
    "outubro": "outubro",
    "out": "outubro",
    "10": "outubro",
    "novembro": "novembro",
    "nov": "novembro",
    "11": "novembro",
    "dezembro": "dezembro",
    "dez": "dezembro",
    "12": "dezembro",
}


def normalizar_texto(texto):
    texto = unicodedata.normalize("NFKD", str(texto))
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return texto.lower()


def detectar_mes_por_nome(nome_arquivo):
    nome = normalizar_texto(Path(nome_arquivo).stem)
    tokens = re.split(r"[^a-z0-9]+", nome)
    for token in tokens:
        if token in MESES_ALIASES:
            return MESES_ALIASES[token]
    return None


def parse_valor_brasileiro(valor):
    if valor in (None, ""):
        return 0.0
    texto = str(valor).strip()
    if not texto:
        return 0.0

    sinal = 1
    texto_upper = texto.upper()
    if texto_upper.endswith("D"):
        sinal = -1
    elif texto_upper.endswith("C"):
        sinal = 1

    texto = texto_upper.removesuffix("D").removesuffix("C").strip()
    texto = texto.replace(".", "").replace(",", ".")
    try:
        return round(sinal * abs(float(texto)), 2)
    except ValueError:
        return 0.0


def coluna_para_indice(coluna):
    indice = 0
    for char in coluna:
        indice = indice * 26 + (ord(char.upper()) - ord("A") + 1)
    return indice - 1


def obter_layout_colunas(input_config, total_colunas, base_zero=True):
    configured = {
        nome: coluna_para_indice(coluna)
        for nome, coluna in input_config["columns"].items()
    }

    # Alguns balancetes exportam sem a coluna "Movimento":
    # A=Descricao, D=Saldo Anterior, E=Debito, F=Credito, G=Saldo Atual.
    if total_colunas <= 7:
        layout = {
            "description": 0,
            "saldo_anterior": 3,
            "debito": 4,
            "credito": 5,
            "saldo_atual": 6,
        }
    else:
        layout = configured

    if not base_zero:
        return {nome: indice + 1 for nome, indice in layout.items()}
    return layout


def calcular_movimento_se_necessario(valores):
    if "movimento" not in valores:
        valores["movimento"] = round(valores.get("saldo_atual", 0.0) - valores.get("saldo_anterior", 0.0), 2)
    return valores


def extrair_balancete_xls(caminho, config):
    caminho = Path(caminho)
    if caminho.suffix.lower() == ".xlsx":
        book = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        return extrair_balancete_openpyxl_workbook(book, config)

    book = xlrd.open_workbook(caminho)
    return extrair_balancete_xlrd_workbook(book, config)


def extrair_balancete_xls_bytes(conteudo, config, filename=None):
    if filename and Path(filename).suffix.lower() == ".xlsx":
        book = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
        return extrair_balancete_openpyxl_workbook(book, config)

    book = xlrd.open_workbook(file_contents=conteudo)
    return extrair_balancete_xlrd_workbook(book, config)


def extrair_balancete_xlrd_workbook(book, config):
    input_config = config["input"]
    sheet = book.sheet_by_name(input_config["sheet"])

    layout = obter_layout_colunas(input_config, sheet.ncols)
    col_descricao = layout["description"]
    colunas_valores = {nome: indice for nome, indice in layout.items() if nome != "description"}

    pattern = re.compile(input_config["classification_pattern"])
    valores = defaultdict(lambda: defaultdict(float))

    for row_idx in range(sheet.nrows):
        descricao = str(sheet.cell_value(row_idx, col_descricao)).strip()
        match = pattern.search(descricao)
        if not match:
            continue

        classificacao = str(match.group(1)).strip()
        valores_linha = {}
        for nome_coluna, col_idx in colunas_valores.items():
            valor_bruto = sheet.cell_value(row_idx, col_idx) if col_idx < sheet.ncols else None
            valores_linha[nome_coluna] = parse_valor_brasileiro(valor_bruto)

        for nome_coluna, valor in calcular_movimento_se_necessario(valores_linha).items():
            valores[classificacao][nome_coluna] = round(valores[classificacao][nome_coluna] + valor, 2)

    return valores


def extrair_balancete_openpyxl_workbook(book, config):
    input_config = config["input"]
    sheet = book[input_config["sheet"]]

    layout = obter_layout_colunas(input_config, sheet.max_column, base_zero=False)
    col_descricao = layout["description"]
    colunas_valores = {nome: indice for nome, indice in layout.items() if nome != "description"}

    pattern = re.compile(input_config["classification_pattern"])
    valores = defaultdict(lambda: defaultdict(float))

    max_col = max([col_descricao, *colunas_valores.values()])

    for row in sheet.iter_rows(max_col=max_col):
        descricao_cell = row[col_descricao - 1] if len(row) >= col_descricao else None
        descricao = str((descricao_cell.value if descricao_cell else None) or "").strip()
        match = pattern.search(descricao)
        if not match:
            continue

        classificacao = str(match.group(1)).strip()
        valores_linha = {}
        for nome_coluna, col_idx in colunas_valores.items():
            cell = row[col_idx - 1] if len(row) >= col_idx else None
            valores_linha[nome_coluna] = parse_valor_brasileiro(cell.value if cell else None)

        for nome_coluna, valor in calcular_movimento_se_necessario(valores_linha).items():
            valores[classificacao][nome_coluna] = round(valores[classificacao][nome_coluna] + valor, 2)

    return valores


def classificacao_casa(classificacao, regra):
    for alvo in regra["classifications"]:
        if regra["match"] == "exact" and classificacao == alvo:
            return True
        if regra["match"] == "prefix" and (
            classificacao == alvo or classificacao.startswith(f"{alvo}.")
        ):
            return True
    return False


def somar_regra(valores_balancete, regra):
    coluna = regra["value_column"]
    alvos = regra["classifications"]

    if regra["match"] == "prefix":
        # Usa o totalizador pai se ele existir (ex: [5] ou [7]).
        # O totalizador é calculado pelo sistema contábil e bate exatamente
        # com a soma das folhas. Somar todos os filhos causaria duplicação
        # porque o balancete inclui níveis intermediários (5, 5.2, 5.2.01, ...).
        total_exato = 0.0
        encontrou_exato = False
        for alvo in alvos:
            if alvo in valores_balancete:
                encontrou_exato = True
                total_exato = round(total_exato + valores_balancete[alvo].get(coluna, 0.0), 2)
        if encontrou_exato:
            return total_exato

        # Totalizador não existe: soma apenas as contas FOLHA (sem filhos)
        # para evitar dupla contagem dos níveis intermediários.
        todas = set(valores_balancete.keys())
        total = 0.0
        for classificacao, valores in valores_balancete.items():
            if classificacao in alvos:
                continue  # pula o próprio totalizador se vier a existir
            if not classificacao_casa(classificacao, regra):
                continue
            # É folha se nenhuma outra classificação começa com "classificacao."
            tem_filho = any(
                outro != classificacao and outro.startswith(f"{classificacao}.")
                for outro in todas
            )
            if not tem_filho:
                total = round(total + valores.get(coluna, 0.0), 2)
        return total

    # match == "exact": soma somente as contas listadas explicitamente
    total = 0.0
    for classificacao, valores in valores_balancete.items():
        if classificacao_casa(classificacao, regra):
            total = round(total + valores.get(coluna, 0.0), 2)
    return total


def _gravar_celula(sheet, referencia, valor):
    """Grava valor na célula (sobrescreve mesmo se contiver fórmula no modelo)."""
    cell = sheet[referencia]
    cell.value = valor


def aplicar_bi_resultados(sheet, modelo_config, mes, trimestre, bi_valores):
    """Aplica valores pré-calculados do BI Fiscal nas células da planilha modelo."""
    bi_mappings = modelo_config.get("bi_cfop_mappings", [])
    if not bi_mappings or not bi_valores:
        return

    if "month_columns_by_quarter" in modelo_config:
        coluna = modelo_config["month_columns_by_quarter"][trimestre][mes]
        for regra in bi_mappings:
            valor = bi_valores.get(regra["id"], 0.0)
            _gravar_celula(sheet, f"{coluna}{regra['rows_by_quarter'][trimestre]}", valor)
        return

    coluna = modelo_config["month_columns"][mes]
    for regra in bi_mappings:
        valor = bi_valores.get(regra["id"], 0.0)
        _gravar_celula(sheet, f"{coluna}{regra['row']}", valor)


def aplicar_resultados(sheet, modelo_config, mes, trimestre, valores_balancete):
    if "month_columns" in modelo_config:
        coluna = modelo_config["month_columns"][mes]
        for regra in modelo_config["result_mappings"]:
            _gravar_celula(sheet, f"{coluna}{regra['row']}", somar_regra(valores_balancete, regra))
        return

    coluna = modelo_config["month_columns_by_quarter"][trimestre][mes]
    for regra in modelo_config["result_mappings"]:
        linha = regra["rows_by_quarter"][trimestre]
        _gravar_celula(sheet, f"{coluna}{linha}", somar_regra(valores_balancete, regra))


def aplicar_adicoes(workbook, modelo_config, mes, trimestre, valores_balancete):
    if "additions_sheet" not in modelo_config:
        return

    sheet = workbook[modelo_config["additions_sheet"]]
    regras = modelo_config.get("additions_mappings", [])
    if not regras:
        return

    if "additions_month_columns" in modelo_config:
        coluna = modelo_config["additions_month_columns"][mes]
        for regra in regras:
            _gravar_celula(sheet, f"{coluna}{regra['row']}", somar_regra(valores_balancete, regra))
        return

    coluna = modelo_config["additions_month_columns_by_quarter"][trimestre][mes]
    for regra in regras:
        linha = regra["rows_by_quarter"][trimestre]
        _gravar_celula(sheet, f"{coluna}{linha}", somar_regra(valores_balancete, regra))


def processar(config_path, modelo, balancete_path, output_path, mes=None, nome=None, cnpj=None):
    with open(config_path, "r", encoding="utf-8") as arquivo:
        config = json.load(arquivo)

    mes = mes or detectar_mes_por_nome(balancete_path)
    if not mes:
        raise ValueError("Mes nao identificado. Informe --month ou coloque o mes no nome do arquivo.")
    mes = MESES_ALIASES[normalizar_texto(mes)]
    trimestre = config["months"][mes]["quarter"]

    modelo_config = config["models"][modelo]
    valores_balancete = extrair_balancete_xls(balancete_path, config)

    workbook = openpyxl.load_workbook(modelo_config["template_file"])
    sheet_resultado = workbook[modelo_config["result_sheet"]]

    if nome:
        sheet_resultado[modelo_config["company_cells"]["nome"]] = nome
    if cnpj:
        sheet_resultado[modelo_config["company_cells"]["cnpj"]] = cnpj

    aplicar_resultados(sheet_resultado, modelo_config, mes, trimestre, valores_balancete)
    aplicar_adicoes(workbook, modelo_config, mes, trimestre, valores_balancete)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return {
        "modelo": modelo,
        "mes": mes,
        "trimestre": trimestre,
        "classificacoes_lidas": len(valores_balancete),
        "arquivo_saida": str(output_path),
    }


def main():
    parser = argparse.ArgumentParser(description="Processa balancete no novo formato.")
    parser.add_argument("--config", default="mapeamentos_novo_formato.json")
    parser.add_argument("--model", choices=["lucro_presumido", "lucro_real", "mensal"], required=True)
    parser.add_argument("--balancete", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--month")
    parser.add_argument("--nome")
    parser.add_argument("--cnpj")
    args = parser.parse_args()

    resumo = processar(
        config_path=args.config,
        modelo=args.model,
        balancete_path=args.balancete,
        output_path=args.output,
        mes=args.month,
        nome=args.nome,
        cnpj=args.cnpj,
    )
    print(json.dumps(resumo, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
